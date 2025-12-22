# Import Library
import sys, os, csv
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import firebase_admin
from firebase_admin import credentials, firestore

from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtCore import (QCoreApplication, QPropertyAnimation, QDate, QDateTime, QMetaObject, QObject, QPoint, QRect, QSize, QTime, QTimer, QUrl, Qt, QEvent, QStandardPaths)

from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor, QFont, QFontDatabase, QIcon, QKeySequence, QLinearGradient, QPalette, QPainter, QPixmap, QRadialGradient)
from PySide6.QtWidgets import *

from PySide6.QtGui import QIntValidator, QDoubleValidator

# Mengimport GUI File
from ui_mainwindow import Ui_MainWindow
from ui_functions import WindowFunctions

# Class untuk mengatur Hari dan Waktu pada MainWindow dan InfoWidget
class Date():
    def update_time(self, label):
        # Ambil data waktu dan tanggal sekarang.
        current_time = QDateTime.currentDateTime()

        # Format waktu dan tanggal menjadi "HH:mm - dddd, dd MMMM yyyy"
        time_text = current_time.toString("HH:mm")
        date_text = current_time.toString("dddd, dd MMMM yyyy")

        # Menggabungkan waktu dan tanggal into one string
        full_text = f"{time_text} - {date_text}"

        # Mengupdate label pada MainWindow dan Widget dengan QCoreApplication.translate
        label.setText(QCoreApplication.translate("MainWindow", full_text, None))

class MainWindow(QMainWindow,Date):
    def __init__(self):
        super().__init__()

        # Inisialisasi tampilan GUI utama
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        
        # Method untuk inisialisasi GUI yang saat aplikasi berjalan.
        self.initUI()
    
        # Integrasikan WindowFunctions
        self.window_functions = WindowFunctions(self, self.ui.drop_shadow_frame)

        # Mengatur tombol minimize, maximize/restore, dan close
        self.ui.btn_maximize.clicked.connect(lambda: self.window_functions.toggle_maximize_restore(self.ui.btn_maximize))
        self.ui.btn_minimize.clicked.connect(self.showMinimized)
        self.ui.btn_close.clicked.connect(self.close)
        
        # Mengatur waktu untuk mengupdate waktu dan tanggal pada label.
        self.timer = QTimer(self)
        self.timer.timeout.connect(lambda: self.update_time(self.ui.date_label)) 
        self.timer.start(1000)  # Update waktu setiap 1 detik.

        # Firebase Setup
        self.setup_firebase()
        
        # Timer untuk mengupdate data sensor secara berkala (misal setiap 5 detik)
        self.sensor_timer = QTimer(self)
        self.sensor_timer.timeout.connect(self.update_sensor_data)
        self.sensor_timer.start(5000)  # Update data sensor setiap 5 detik
        
        # Inisialisasi mode (True untuk Barang Masuk, False untuk Barang Keluar)
        self.mode = True  # Default: Barang Masuk
        
        # Hubungkan tombol info untuk mengubah mode
        self.ui.btn_info.clicked.connect(self.toggle_mode)
        
        # Validasi untuk input Jumlah (hanya angka bulat)
        quantity_validator = QIntValidator(0, 1000000000)  # Batas bawah 0, batas atas 1000000000
        self.ui.lineEdit_jumlah.setValidator(quantity_validator)

        # Hubungkan tombol add untuk menambahkan data yang diinputkan user ke database.
        self.ui.btn_add.clicked.connect(self.read_user_input)
        # Hubungkan tombol remove untuk menghapus data pada database.
        self.ui.btn_remove.clicked.connect(self.btn_remove_clicked)    
        # Hubungkan tombol download untuk mendownload tabel.
        self.ui.btn_download.clicked.connect(self.download_table_to_excel)
        
        # Menampilkan jendela dalam mode normal
        self.show()
        
        # Menampilkan Data dari Database ke Tabel
        self.load_table_data()
        
        # Cek Penghapusan Database
        self.check_date_and_clear_database_once() #debug_date=QDate(2025, 3, 1) 
    
    def initUI(self):
        """Pengaturan Awal GUI"""
        # Mengatur Judul Aplikasi
        self.setWindowTitle("Warehouse Information System")
    
        # Mengatur Icon Aplikasi
        pixmap = QPixmap(self.resource_path("images\\icon.ico"))
        icon = QIcon(pixmap.scaled(64, 64, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        self.setWindowIcon(icon)
        
        self.update_table_margin()
        self.ui.tabeldata.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.ui.lineEdit_unused.setEnabled(False)
        self.ui.dateEdit.setDate(QtCore.QDate.currentDate())
        
    def update_table_margin(self):
        """Update layoutBottomMargin berdasarkan tinggi aplikasi."""
        height = self.height()

        if height <= 768:
            bottom_margin = 165
        elif height >= 1080:
            bottom_margin = 30
        else:
            # Interpolasi linear antara 736p (160) dan 1080p (30)
            bottom_margin = int(160 - ((height - 736) / (1080 - 736) * (160 - 30)))

        self.ui.horizontalLayout_6.setContentsMargins(30, 20, 30, bottom_margin)
        
    def resizeEvent(self, event):
        """Event yang dipanggil saat ukuran window berubah."""
        self.update_table_margin()
        super().resizeEvent(event)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.dragPos = event.globalPosition().toPoint()

    def toggle_mode(self):
        """
        Berpindah antara mode Barang Masuk dan Barang Keluar.
        """
        # Ubah label sesuai mode
        if self.mode:
            self.ui.status_label_barang.setText("Barang Keluar")
            self.ui.status_label_barang.setStyleSheet(u"QLabel {\n"
            "    background-color: rgb(220, 50, 50); /* Warna background baru */\n"
            "    color: white;  /* Warna teks */\n"
            "    padding: 2px;  /* Padding untuk memberi jarak antara teks dan border */\n"
            "    border-radius: 5px;  /* Membuat sudut bulat\n"
            "}")
            self.ui.label_tanggal.setText("Tanggal Keluar")
            self.ui.label_jumlah.setText("Jumlah Keluar")
            self.ui.tabeldata.horizontalHeaderItem(1).setText("Tanggal Keluar")
            self.ui.tabeldata.horizontalHeaderItem(4).setText("Jumlah Keluar")
            self.clear_inputs()

        else:
            self.ui.status_label_barang.setText("Barang Masuk")
            self.ui.status_label_barang.setStyleSheet(u"QLabel {\n"
            "    background-color: rgb(71, 183, 92); /* Warna background baru */\n"
            "    color: white;  /* Warna teks */\n"
            "    padding: 2px;  /* Padding untuk memberi jarak antara teks dan border */\n"
            "    border-radius: 5px;  /* Membuat sudut bulat\n"
            "}")
            self.ui.label_tanggal.setText("Tanggal Masuk")
            self.ui.label_jumlah.setText("Jumlah Masuk")
            self.ui.tabeldata.horizontalHeaderItem(1).setText("Tanggal Masuk")
            self.ui.tabeldata.horizontalHeaderItem(4).setText("Jumlah Masuk")
            self.clear_inputs()

        # Toggle mode
        self.mode = not self.mode

        # Bersihkan tabel dan muat data sesuai mode baru
        self.ui.tabeldata.setRowCount(0)  # Kosongkan tabel
        self.load_table_data() # Load Tabel
        

    def clear_inputs(self):
        """
        Mengosongkan semua input setelah data disimpan.
        """
        self.ui.lineEdit_merek.clear()
        self.ui.dateEdit.setDate(QtCore.QDate.currentDate())
        self.ui.lineEdit_namabarang.clear()
        self.ui.lineEdit_jumlah.clear()
        self.ui.lineEdit_rak.clear()
        
    def format_angka(self, number):
        """
        Mengubah format angka 1000 ke 1.000,dll
        """
        return f"{int(number):,}".replace(",", ".")
    
    def setup_firebase(self):
        """
        Method untuk setup Firebase.
        """
        # Inisialisasi Firebase Admin SDK
        cred = credentials.Certificate(self.resource_path("servicefiles/serviceAccountKey.json"))
        firebase_admin.initialize_app(cred)

        # Referensi ke database Firestore
        self.db = firestore.client()

    def update_sensor_data(self):
        """
        Method untuk mengambil data pembacaan sensor dari database.
        """
        # Referensi ke dokumen sensor_readings di koleksi DataSensor
        doc_ref = self.db.collection('DataSensor').document('sensor_readings')

        # Mengambil data dari dokumen
        doc = doc_ref.get()

        if doc.exists:
            data = doc.to_dict()
            # Memperbarui label dengan data yang diambil
            humidity = data.get('humidity', 'N/A')
            suhu = data.get('suhu', 'N/A')
            light = data.get('tingkat_cahaya', 'N/A')

            self.ui.label_humid.setText(QCoreApplication.translate("MainWindow", f"Kelembaban: {humidity}%", None))
            self.ui.label_suhu.setText(QCoreApplication.translate("MainWindow", f"Suhu: {suhu}°C", None))
            self.ui.label_light.setText(QCoreApplication.translate("MainWindow", f"Tingkat Cahaya: {light}%", None))
            self.save_sensor_data_to_csv(suhu, humidity, light)
        else:
            print("Dokumen Sensor tidak ditemukan.")
            
    def save_sensor_data_to_csv(self, suhu, kelembaban, tingkat_cahaya):  
        """
        Method untuk menyimpan data pembacaan sensor ke CSV. 
        """
        # Mendapatkan waktu saat ini sebagai timestamp
        timestamp = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")

        # Format tanggal untuk nama file
        current_date = QDateTime.currentDateTime().toString("ddMMyyyy")

        # Nama file CSV berdasarkan tanggal
        file_name = f"databases/data_sensor/data_sensor_{current_date}.csv"
        
        self.resource_path(file_name)
        # Data yang akan disimpan ke CSV
        data = [timestamp, suhu, kelembaban, tingkat_cahaya]

        # Menyimpan data ke dalam CSV
        with open(file_name, mode='a', newline='') as file:
            writer = csv.writer(file)

            # Menulis header jika file kosong
            if file.tell() == 0:
                writer.writerow(["Timestamp", "Suhu", "Kelembaban", "Tingkat Cahaya"])

            # Menulis data
            writer.writerow(data)
            
    def read_user_input(self):
        """
        Membaca data dari input user dan memperbarui database berdasarkan mode (Barang Masuk/Keluar).
        """
        try:
            # Membaca data dari form
            merek_barang = self.ui.lineEdit_merek.text().strip()
            nama_barang = self.ui.lineEdit_namabarang.text().strip()
            jumlah = int(self.ui.lineEdit_jumlah.text().strip())
            rak = self.ui.lineEdit_rak.text().strip()
            tanggal = self.ui.dateEdit.date().toString("dd-MM-yyyy")
            
            # Validasi apakah semua field sudah diisi
            if not merek_barang or not nama_barang or not jumlah or not rak:
                QtWidgets.QMessageBox.warning(
                    self, "Error", "Harap isi semua kolom input sebelum melanjutkan!"
                )
                return

            # Validasi jumlah harus berupa angka positif
            try:
                jumlahcek = int(jumlah)
                if jumlahcek <= 0:
                    raise ValueError
            except ValueError:
                QtWidgets.QMessageBox.warning(
                    self, "Error", "Kolom jumlah harus berupa angka positif!"
                )
                return
        
            if self.mode:  # Mode Barang Masuk
                # Data untuk koleksi BarangMasuk
                collection_name = 'BarangMasuk'
                data_barang = {
                    'merek_barang': merek_barang,
                    'nama_barang': nama_barang,
                    'jumlah': jumlah,
                    'rak': rak,
                    'tanggal': tanggal,
                    'waktu_input': datetime.now(),
                }
                # Tambahkan data ke koleksi BarangMasuk
                self.db.collection(collection_name).add(data_barang)

                # Update stok barang di koleksi StokBarang
                self.update_stok_barang_masuk(merek_barang, jumlah, nama_barang, rak)
                self.load_table_data()

            else:  # Mode Barang Keluar
                # Data untuk koleksi BarangKeluar
                collection_name = 'BarangKeluar'
                data_barang = {
                    'merek_barang': merek_barang,
                    'nama_barang': nama_barang,
                    'jumlah': jumlah,
                    'rak': rak,
                    'tanggal': tanggal,
                    'waktu_input': datetime.now(),
                }

                # Validasi stok barang di koleksi StokBarang sebelum menambahkan
                if self.update_stok_barang_keluar(merek_barang, jumlah, nama_barang, rak):
                    # Jika stok mencukupi, tambahkan data ke koleksi BarangKeluar
                    self.db.collection(collection_name).add(data_barang)
                else:
                    QtWidgets.QMessageBox.warning(
                        self,
                        "Error",
                        f"Stok barang tidak mencukupi!\n\n"
                        f"Detail:\n"
                        f"Merek Barang   : {merek_barang}\n"
                        f"Nama Barang    : {nama_barang}\n"
                        f"Rak            : {rak}"
                    )
                    return

            # Memberi feedback kepada pengguna
            QtWidgets.QMessageBox.information(self, "Success", f"Data berhasil ditambahkan ke koleksi {collection_name}!")

            # Mengosongkan input setelah data disimpan
            self.clear_inputs()
            self.load_table_data()
        
        except Exception as e:
            QtWidgets.QMessageBox.warning(
                    self, "Error", "Harap isi semua kolom input sebelum melanjutkan!")
 

    def update_stok_barang_masuk(self, merek_barang, jumlah_masuk, nama_barang, rak):
        """
        Memperbarui stok barang di koleksi StokBarang untuk barang masuk.
        Jika ID barang sudah ada, tambahkan jumlah stoknya. Jika tidak ada, buat data baru.
        """
        stok_barang_snapshot = self.db.collection('StokBarang') \
        .where('merek_barang', '==', merek_barang) \
        .where('nama_barang', '==', nama_barang) \
        .where('rak', '==', rak) \
        .get()

        if stok_barang_snapshot:
            # Jika barang sudah ada, update stok
            for doc in stok_barang_snapshot:
                doc_ref = doc.reference
                data = doc.to_dict()
                current_stok = int(data['stok_barang'])
                new_stok = current_stok + jumlah_masuk
                doc_ref.update({'stok_barang': new_stok})
        else:
            # Jika barang belum ada, tambahkan data baru
            self.db.collection('StokBarang').add({
                'merek_barang': merek_barang,
                'nama_barang': nama_barang,
                'stok_barang': jumlah_masuk,
                'rak': rak,
            })


    def update_stok_barang_keluar(self, merek_barang, jumlah_keluar, nama_barang, rak):
        """
        Memperbarui stok barang di koleksi StokBarang untuk barang keluar.
        Jika stok mencukupi, kurangi stoknya. Jika tidak, kembalikan False.
        """
        stok_barang_snapshot = self.db.collection('StokBarang') \
        .where('merek_barang', '==', merek_barang) \
        .where('nama_barang', '==', nama_barang) \
        .where('rak', '==', rak) \
        .get()
        
        if stok_barang_snapshot:
            # Jika barang ada, cek stok dan update
            for doc in stok_barang_snapshot:
                doc_ref = doc.reference
                data = doc.to_dict()
                current_stok = int(data['stok_barang'])

                if current_stok >= jumlah_keluar:
                    # Stok mencukupi, kurangi stok
                    new_stok = current_stok - jumlah_keluar
                    doc_ref.update({'stok_barang': new_stok})
                    return True
                else:
                    # Stok tidak mencukupi
                    return False
        else:
            # Barang tidak ditemukan di StokBarang
            QtWidgets.QMessageBox.warning(self, 
            "Error", 
            f"Barang tidak ditemukan di StokBarang!\n\n"
            f"Detail:\n"
            f"Merek Barang   : {merek_barang}\n"
            f"Nama Barang    : {nama_barang}\n"
            f"Rak            : {rak}"
            )
            return False

    def load_table_data(self):
        """
        Menampilkan data pada QTableWidget sesuai mode (Barang Masuk/Barang Keluar).
        """
        # Bersihkan data tabel sebelumnya
        self.ui.tabeldata.setRowCount(0)
        
        if self.mode:  # Mode Barang Masuk
            collection_name = "BarangMasuk"
            try:
                # data_snapshot = self.db.collection(collection_name).get()
                data_snapshot = self.db.collection(collection_name).order_by('waktu_input', direction=firestore.Query.DESCENDING).get()

                # Tampilkan data di tabel
                for row_number, doc in enumerate(data_snapshot):
                    data = doc.to_dict()
                    self.ui.tabeldata.insertRow(row_number)
                    self.ui.tabeldata.setItem(row_number, 0, QTableWidgetItem(str(row_number + 1)))  # Nomor
                    self.ui.tabeldata.setItem(row_number, 1, QTableWidgetItem(data["tanggal"]))
                    self.ui.tabeldata.setItem(row_number, 2, QTableWidgetItem(data["merek_barang"]))
                    self.ui.tabeldata.setItem(row_number, 3, QTableWidgetItem(data["nama_barang"]))
                    
                    formatted_angka = self.format_angka(data["jumlah"])
                    self.ui.tabeldata.setItem(row_number, 4, QTableWidgetItem(formatted_angka))
                    
                    self.ui.tabeldata.setItem(row_number, 6, QTableWidgetItem(data["rak"]))  # Asal Kirim
                    self.ui.tabeldata.resizeColumnsToContents()
                    self.ui.tabeldata.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                    
                    # Ambil stok dari koleksi StokBarang berdasarkan merek_barang
                    # stok_barang_ref = self.db.collection('StokBarang').where('merek_barang', '==', data["merek_barang"]).get()
                    stok_barang_ref = self.db.collection('StokBarang') \
                    .where('merek_barang', '==', data["merek_barang"]) \
                    .where('nama_barang', '==', data["nama_barang"]) \
                    .where('rak', '==', data["rak"]) \
                    .get()

                    stok_value = "N/A"  # Default value jika stok tidak ditemukan
                    if stok_barang_ref:
                        stok_value = stok_barang_ref[0].to_dict().get('stok_barang', "N/A")
                    
                    formatted_stok = self.format_angka(stok_value)
                    self.ui.tabeldata.setItem(row_number, 5, QTableWidgetItem(formatted_stok))
            except Exception as e:
                print(f"Error loading Barang Masuk data: {e}")

        else:  # Mode Barang Keluar
            collection_name = "BarangKeluar"
            try:
                # data_snapshot = self.db.collection().get()
                data_snapshot = self.db.collection(collection_name).order_by('waktu_input', direction=firestore.Query.DESCENDING).get()

                # Tampilkan data di tabel
                for row_number, doc in enumerate(data_snapshot):
                    data = doc.to_dict()
                    self.ui.tabeldata.insertRow(row_number)
                    self.ui.tabeldata.setItem(row_number, 0, QTableWidgetItem(str(row_number + 1)))  # Nomor
                    self.ui.tabeldata.setItem(row_number, 1, QTableWidgetItem(data["tanggal"]))
                    self.ui.tabeldata.setItem(row_number, 2, QTableWidgetItem(data["merek_barang"]))
                    self.ui.tabeldata.setItem(row_number, 3, QTableWidgetItem(data["nama_barang"]))
                    
                    formatted_angka = self.format_angka(data["jumlah"])
                    self.ui.tabeldata.setItem(row_number, 4, QTableWidgetItem(formatted_angka))
                    
                    self.ui.tabeldata.setItem(row_number, 6, QTableWidgetItem(data["rak"]))  # rak_out

                    # Ambil stok dari koleksi StokBarang berdasarkan merek_barang
                    # stok_barang_ref = self.db.collection('StokBarang').where('merek_barang', '==', data["merek_barang"]).get()
                    stok_barang_ref = self.db.collection('StokBarang') \
                    .where('merek_barang', '==', data["merek_barang"]) \
                    .where('nama_barang', '==', data["nama_barang"]) \
                    .where('rak', '==', data["rak"]) \
                    .get()

                    stok_value = "N/A"  # Default value jika stok tidak ditemukan
                    if stok_barang_ref:
                        stok_value = stok_barang_ref[0].to_dict().get('stok_barang', "N/A")
                    
                    formatted_stok = self.format_angka(stok_value)
                    self.ui.tabeldata.setItem(row_number, 5, QTableWidgetItem(formatted_stok))
                    # self.ui.tabeldata.setItem(row_number, 7, QTableWidgetItem(str(stok_value)))  # Stok
                    self.ui.tabeldata.resizeColumnsToContents()
                    self.ui.tabeldata.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            except Exception as e:
                print(f"Error loading Barang Keluar data: {e}")
                
    def btn_remove_clicked(self):
        """
        Menghapus baris data yang dipilih sesuai mode (BarangMasuk atau BarangKeluar).
        """
        selected_row = self.ui.tabeldata.currentRow()
        if selected_row == -1:  # Jika tidak ada baris yang dipilih
            QtWidgets.QMessageBox.warning(self, "Error", "Pilih baris yang ingin dihapus terlebih dahulu.")
            return

        tanggal = self.ui.tabeldata.item(selected_row, 1).text()
        merek_barang = self.ui.tabeldata.item(selected_row, 2).text()
        nama_barang = self.ui.tabeldata.item(selected_row, 3).text()
        rak = self.ui.tabeldata.item(selected_row, 6).text()

        try:
            if self.mode:  # Mode BarangMasuk
                self.remove_barang_masuk(merek_barang, tanggal, selected_row, nama_barang, rak)
            else:  # Mode BarangKeluar
                self.remove_barang_keluar(merek_barang, tanggal, selected_row, nama_barang, rak)
        except Exception as e:
            print(f"Error saat menghapus data: {e}")
            QtWidgets.QMessageBox.critical(self, "Error", "Terjadi kesalahan saat menghapus data.")

    def remove_barang_masuk(self, merek_barang, tanggal, selected_row, nama_barang, rak):
        """
        Menghapus data tertentu dari BarangMasuk dengan validasi stok.
        """
        try:
            # Ambil jumlah yang akan dihapus dari tabel
            jumlah_dihapus = int(self.ui.tabeldata.item(selected_row, 4).text().replace(".", ""))

            # Hitung stok sebelum penghapusan
            total_masuk = self.get_total_barang_masuk(merek_barang, nama_barang, rak)
            total_keluar = self.get_total_barang_keluar(merek_barang, nama_barang, rak)
            stok_sekarang = total_masuk - total_keluar

            # Cek apakah stok akan negatif setelah penghapusan
            if stok_sekarang - jumlah_dihapus < 0:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Error",
                    f"Gagal menghapus Data! \n\n"
                    f"Detail:\n"
                    f"- Merek: {merek_barang}\n"
                    f"- Nama Barang: {nama_barang}\n"
                    f"- Rak: {rak}\n\n"
                    f"Stok tidak mencukupi untuk penghapusan ini.\n"
                    f"Solusi:\n"
                    f"- Hapus data barang keluar (sebelumnya) terlebih dahulu agar penghapusan barang masuk ini dapat dilakukan."
                ) 
                return

            # Cari dokumen BarangMasuk yang akan dihapus
            data_snapshot = self.db.collection('BarangMasuk') \
                .where('merek_barang', '==', merek_barang) \
                .where('nama_barang', '==', nama_barang) \
                .where('tanggal', '==', tanggal) \
                .where('jumlah', '==', jumlah_dihapus) \
                .where('rak', '==', rak) \
                .get()

            if not data_snapshot:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Error",
                    f"Tidak ada data ditemukan!\n\n"
                    f"Detail:\n"
                    f"- Merek: {merek_barang}\n"
                    f"- Nama Barang: {nama_barang}\n"
                    f"- Rak: {rak}\n"
                    f"- Tanggal: {tanggal}\n"
                    f"- Jumlah: {jumlah_dihapus}\n\n"
                    f"Silakan periksa kembali data yang Anda masukkan."
                )
                return

            # Hapus hanya satu dokumen pertama yang sesuai
            for doc in data_snapshot:
                doc.reference.delete()
                break

            # Update stok setelah penghapusan
            self.update_stok_barang_masuk(merek_barang, -jumlah_dihapus, nama_barang, rak)

            # Hapus baris dari tabel
            self.ui.tabeldata.removeRow(selected_row)

            # Periksa stok barang apakah perlu dihapus
            self.remove_stok_barang(merek_barang, nama_barang, rak)

            # Periksa apakah masih ada data BarangMasuk untuk barang ini
            barang_masuk_snapshot = self.db.collection('BarangMasuk') \
                .where('merek_barang', '==', merek_barang) \
                .where('nama_barang', '==', nama_barang) \
                .where('rak', '==', rak) \
                .get()

            if not barang_masuk_snapshot:  # Jika tidak ada data BarangMasuk yang tersisa
                print(f"Tidak ada lagi data BarangMasuk untuk Merek: {merek_barang} . Menghapus data terkait...")

                # Hapus semua data di BarangKeluar terkait barang ini
                barang_keluar_snapshot = self.db.collection('BarangKeluar') \
                    .where('merek_barang', '==', merek_barang) \
                    .where('nama_barang', '==', nama_barang) \
                    .where('rak', '==', rak) \
                    .get()
                for doc in barang_keluar_snapshot:
                    doc.reference.delete()

                # Hapus data di StokBarang
                stok_barang_snapshot = self.db.collection('StokBarang') \
                    .where('merek_barang', '==', merek_barang) \
                    .where('nama_barang', '==', nama_barang) \
                    .where('rak', '==', rak) \
                    .get()
                for doc in stok_barang_snapshot:
                    doc.reference.delete()

            self.load_table_data()

            QtWidgets.QMessageBox.information(
                self,
                "Success",
                f"Data Barang Masuk berhasil dihapus!\n\n"
                f"Detail:\n"
                f"Merek Barang   : {merek_barang}\n"
                f"Nama Barang    : {nama_barang}\n"
                f"Rak            : {rak}"
            )

        except Exception as e:
            print(f"Error saat menghapus data Barang Masuk: {e}")
            QtWidgets.QMessageBox.critical(self, "Error", "Terjadi kesalahan saat menghapus data.")

    def get_total_barang_masuk(self, merek_barang, nama_barang, rak):
        """
        Menghitung total barang masuk berdasarkan merek dan nama barang.
        """
        try:
            barang_masuk_snapshot = self.db.collection('BarangMasuk') \
                .where('merek_barang', '==', merek_barang) \
                .where('nama_barang', '==', nama_barang) \
                .where('rak', '==', rak) \
                .get()

            total_masuk = sum(int(doc.to_dict().get('jumlah', 0)) for doc in barang_masuk_snapshot)
            return total_masuk
        except Exception as e:
            print(f"Error saat menghitung total barang masuk: {e}")
            return 0


    def get_total_barang_keluar(self, merek_barang, nama_barang, rak):
        """
        Menghitung total barang keluar berdasarkan merek dan nama barang.
        """
        try:
            barang_keluar_snapshot = self.db.collection('BarangKeluar') \
                .where('merek_barang', '==', merek_barang) \
                .where('nama_barang', '==', nama_barang) \
                .where('rak', '==', rak) \
                .get()

            total_keluar = sum(int(doc.to_dict().get('jumlah', 0)) for doc in barang_keluar_snapshot)
            return total_keluar
        except Exception as e:
            print(f"Error saat menghitung total barang keluar: {e}")
            return 0


    def remove_barang_keluar(self, merek_barang, tanggal, selected_row, nama_barang, rak):
        """
        Menghapus data tertentu dari BarangKeluar berdasarkan ID, tanggal, jumlah, dan waktu_input
        dengan mendapatkan waktu_input langsung dari database.
        """
        try:
            # Ambil jumlah yang akan dihapus dari tabel
            jumlah_dihapus = int(self.ui.tabeldata.item(selected_row, 4).text().replace(".", ""))  # Jumlah barang yang dihapus

            # Cari dokumen spesifik di BarangKeluar
            data_snapshot = self.db.collection('BarangKeluar') \
                .where('merek_barang', '==', merek_barang) \
                .where('nama_barang', '==', nama_barang) \
                .where('tanggal', '==', tanggal) \
                .where('jumlah', '==', jumlah_dihapus) \
                .where('rak', '==', rak) \
                .get()

            if not data_snapshot:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Error",
                    f"Tidak ada data ditemukan!\n\n"
                    f"Detail:\n"
                    f"- Merek: {merek_barang}\n"
                    f"- Nama Barang: {nama_barang}\n"
                    f"- Rak: {rak}\n"
                    f"- Tanggal: {tanggal}\n"
                    f"- Jumlah: {jumlah_dihapus}\n\n"
                    f"Silakan periksa kembali data yang Anda masukkan."
                )
                return

            # Hapus hanya dokumen pertama yang sesuai
            for doc in data_snapshot:
                waktu_input = doc.to_dict().get('waktu_input')  # Ambil waktu_input dari dokumen
                print(f"Menghapus dokumen dengan waktu_input: {waktu_input}")  # Debugging (opsional)
                doc.reference.delete()
                break  # Hentikan setelah satu dokumen dihapus

            # Kembalikan stok barang sesuai jumlah yang dihapus
            self.update_stok_barang_keluar(merek_barang, -jumlah_dihapus, nama_barang, rak)

            # Hapus baris dari tabel
            self.ui.tabeldata.removeRow(selected_row)

            # Periksa stok barang apakah perlu dihapus
            self.remove_stok_barang(merek_barang, nama_barang, rak)
            
            self.load_table_data()

            QtWidgets.QMessageBox.information(
                self,
                "Success",
                f"Data Barang Keluar berhasil dihapus!\n\n"
                f"Detail:\n"
                f"- Merek: {merek_barang}\n"
                f"- Nama Barang: {nama_barang}\n"
                f"- Rak: {rak}\n\n"
                f"Data telah dihapus dari sistem."
            )
        except Exception as e:
            print(f"Error saat menghapus data Barang Keluar: {e}")
            QtWidgets.QMessageBox.critical(self, "Error", "Terjadi kesalahan saat menghapus data.")

    def remove_stok_barang(self, merek_barang,nama_barang, rak):
        """
        Mengecek apakah stok barang perlu dihapus jika tidak ada data di BarangMasuk atau BarangKeluar.
        """
        try:
            barang_masuk_exists = self.db.collection('BarangMasuk') \
                .where('merek_barang', '==', merek_barang) \
                .where('nama_barang', '==', nama_barang) \
                .where('rak', '==', rak) \
                .get()
            barang_keluar_exists = self.db.collection('BarangKeluar') \
                .where('merek_barang', '==', merek_barang) \
                .where('nama_barang', '==', nama_barang) \
                .where('rak', '==', rak) \
                .get()

            if not barang_masuk_exists and not barang_keluar_exists:
                # Hapus data stok jika tidak ada data terkait
                stok_barang_snapshot = self.db.collection('StokBarang') \
                .where('merek_barang', '==', merek_barang) \
                .where('nama_barang', '==', nama_barang) \
                .where('rak', '==', rak) \
                .get()
                for doc in stok_barang_snapshot:
                    doc.reference.delete()
                print(f"Stok barang dengan Merek {merek_barang} berhasil dihapus.")
            else:
                print(f"Barang dengan Merek {merek_barang} masih memiliki data terkait, stok tidak dihapus.")
        except Exception as e:
            print(f"Error saat menghapus stok barang: {e}")
        
    def resource_path(self, relative_path):
        """ Mengonversi path relatif menjadi path absolut. 
        Berguna untuk memastikan file dapat ditemukan dari 
        direktori aplikasi saat ini.
        """
        base_path = os.path.abspath(".")  # Mengatur ke directory saat ini. 
        return os.path.join(base_path, relative_path)
    
    def download_table_to_excel(self):
        """
        Menyimpan data dari tabel Barang Masuk dan Barang Keluar ke file Excel.
        Barang Masuk di sebelah kiri, Barang Keluar di sebelah kanan.
        """
        reply = QMessageBox.question(
        self,
        "Konfirmasi",
        "Apakah Anda ingin menyimpan Data Informasi Gudang ke file Excel?",
        QMessageBox.Yes | QMessageBox.No,
        QMessageBox.No
        )

        # Jika pengguna memilih "No", batalkan proses
        if reply == QMessageBox.No:
            return
        
        try:
            # Pilih lokasi penyimpanan file
            options = QFileDialog.Options()
            # Saran nama file berdasarkan tanggal
            default_filename = f"Data_Informasi_Gudang_{datetime.now().strftime('%d-%m-%Y')}.xlsx"

            # Dialog untuk memilih lokasi penyimpanan dengan nama file yang disarankan
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Simpan File Excel", 
                default_filename,  # Nama file default
                "Excel Files (*.xlsx);;All Files (*)", 
                options=options
            )
            
            if not file_path:
                return  # User batal memilih file

            # Ambil data Barang Masuk
            barang_masuk_snapshot = self.db.collection("BarangMasuk").order_by('waktu_input', direction=firestore.Query.DESCENDING).get()

            data_barang_masuk = []
            for idx, doc in enumerate(barang_masuk_snapshot):
                data = doc.to_dict()
                
                # Ambil stok dari koleksi StokBarang berdasarkan merek_barang, nama_barang, dan rak
                stok_barang_ref = self.db.collection('StokBarang') \
                    .where('merek_barang', '==', data["merek_barang"]) \
                    .where('nama_barang', '==', data["nama_barang"]) \
                    .where('rak', '==', data["rak"]) \
                    .get()

                # Default nilai stok jika tidak ditemukan
                stok_value = "N/A"
                if stok_barang_ref:
                    stok_value = stok_barang_ref[0].to_dict().get('stok_barang', "N/A")
                
                # Tambahkan data ke dalam daftar
                data_barang_masuk.append([
                    idx + 1,  # No
                    data["tanggal"],  # Tanggal
                    data["merek_barang"],  # Merek
                    data["nama_barang"],  # Nama Barang
                    data["jumlah"],  # Jumlah
                    stok_value,  # Stok
                    data["rak"]  # Rak
                ])

            # Konversi ke DataFrame
            df_barang_masuk = pd.DataFrame(data_barang_masuk, columns=["No", "Tanggal Masuk", "Merek", "Nama Barang", "Jumlah Masuk", "Stok", "Rak"])
            
            # Ambil data Barang Keluar
            barang_keluar_snapshot = self.db.collection("BarangKeluar").order_by('waktu_input', direction=firestore.Query.DESCENDING).get()

            data_barang_keluar = []
            for idx, doc in enumerate(barang_keluar_snapshot):
                data = doc.to_dict()
                
                # Ambil stok dari koleksi StokBarang berdasarkan merek_barang, nama_barang, dan rak
                stok_barang_ref = self.db.collection('StokBarang') \
                    .where('merek_barang', '==', data["merek_barang"]) \
                    .where('nama_barang', '==', data["nama_barang"]) \
                    .where('rak', '==', data["rak"]) \
                    .get()

                # Default nilai stok jika tidak ditemukan
                stok_value = "N/A"
                if stok_barang_ref:
                    stok_value = stok_barang_ref[0].to_dict().get('stok_barang', "N/A")
                
                # Tambahkan data ke dalam daftar
                data_barang_keluar.append([
                    idx + 1,  # No
                    data["tanggal"],  # Tanggal
                    data["merek_barang"],  # Merek
                    data["nama_barang"],  # Nama Barang
                    data["jumlah"],  # Jumlah
                    stok_value,  # Stok
                    data["rak"]  # Rak
                ])

            # Konversi ke DataFrame
            df_barang_keluar = pd.DataFrame(data_barang_keluar, columns=["No", "Tanggal Keluar", "Merek", "Nama Barang", "Jumlah Keluar", "Stok", "Rak"])

            # Buat file Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Barang"

            # Tambahkan judul utama
            ws.append(["Data Informasi Gudang"])
            ws.append([f"Waktu Download: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            ws.append([])  # Baris kosong

            # Posisi kolom Barang Masuk dan Barang Keluar
            start_col_masuk = 1  # Kolom A
            start_col_keluar = 9  # Kolom H

            # Tambahkan border
            thin_border = Border(
                left=Side(style="thin"), 
                right=Side(style="thin"), 
                top=Side(style="thin"), 
                bottom=Side(style="thin")
            )

            # Tambahkan judul tengah untuk masing-masing tabel
            ws.merge_cells(start_row=4, start_column=start_col_masuk, end_row=4, end_column=start_col_masuk + 6)
            ws.merge_cells(start_row=4, start_column=start_col_keluar, end_row=4, end_column=start_col_keluar + 6)

            cell_masuk = ws.cell(row=4, column=start_col_masuk, value="Barang Masuk")
            cell_keluar = ws.cell(row=4, column=start_col_keluar, value="Barang Keluar")

            # Format judul tengah
            for col in range(start_col_masuk, start_col_masuk + 7):
                ws.cell(row=4, column=col).border = thin_border  # Tambah border ke setiap sel dalam merge

            for col in range(start_col_keluar, start_col_keluar + 7):
                ws.cell(row=4, column=col).border = thin_border

            for cell in [cell_masuk, cell_keluar]:
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Tambahkan header tabel
            barang_masuk_headers = ["No", "Tanggal Masuk", "Merek", "Nama Barang", "Jumlah Masuk", "Stok", "Rak"]
            barang_keluar_headers = ["No", "Tanggal Keluar", "Merek", "Nama Barang", "Jumlah Keluar", "Stok", "Rak"]

            ws.append(barang_masuk_headers + [""] + barang_keluar_headers)

            # Isi data ke Excel
            max_rows = max(len(df_barang_masuk), len(df_barang_keluar))

            for i in range(max_rows):
                row_data = []

                # Barang Masuk
                if i < len(df_barang_masuk):
                    row_data.extend(df_barang_masuk.iloc[i].tolist())
                else:
                    row_data.extend([""] * len(barang_masuk_headers))

                row_data.append("")  # Spasi antar tabel

                # Barang Keluar
                if i < len(df_barang_keluar):
                    row_data.extend(df_barang_keluar.iloc[i].tolist())
                else:
                    row_data.extend([""] * len(barang_keluar_headers))

                ws.append(row_data)

            # Styling untuk judul utama
            ws["A1"].font = Font(size=14, bold=True)
            ws["A2"].font = Font(size=11, italic=True)

            # Styling untuk header tabel
            for cell in ws[5]:  # Header "No, Tanggal, Merek, Nama Barang, Jumlah, Rak"
                cell.font = Font(bold=True)

            # Auto adjust lebar kolom
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].auto_size = True
                
            # Tambahkan border ke seluruh tabel Barang Masuk
            for row in ws.iter_rows(min_row=5, max_row=5 + len(df_barang_masuk), min_col=start_col_masuk, max_col=start_col_masuk + len(barang_masuk_headers) - 1):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

            # Tambahkan border ke seluruh tabel Barang Keluar
            for row in ws.iter_rows(min_row=5, max_row=5 + len(df_barang_keluar), min_col=start_col_keluar, max_col=start_col_keluar + len(barang_keluar_headers) - 1):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    
            # Simpan file Excel
            wb.save(file_path)

            QMessageBox.information(self, "Sukses", "Data berhasil disimpan dalam Excel.")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Terjadi kesalahan saat menyimpan data: {e}, Tutup File Excel yang sedang dibuka!")
            
    def delete_collection(self, collection_name):
        """Menghapus semua dokumen dalam koleksi tertentu."""
        try:
            # Referensi ke koleksi
            collection_ref = self.db.collection(collection_name)
            
            # Mendapatkan semua dokumen dalam koleksi
            docs = collection_ref.get()
            
            # Iterasi untuk menghapus dokumen
            for doc in docs:
                doc.reference.delete()
            
            print(f"Koleksi '{collection_name}' berhasil dihapus.")
        except Exception as e:
            print(f"Terjadi kesalahan saat menghapus koleksi '{collection_name}': {e}")
    
    def delete_all_collections(self):
        """Menghapus isi koleksi StokBarang, BarangMasuk, dan BarangKeluar."""
        collections = ["StokBarang", "BarangMasuk", "BarangKeluar"]
        for collection in collections:
            self.delete_collection(collection)
            
    def show_clear_database_dialog(self):
        """Munculkan dialog konfirmasi sebelum mengosongkan database."""
        # Dialog pertama: Konfirmasi untuk menyimpan data
        dialog = QMessageBox()
        dialog.setWindowTitle("Konfirmasi Pengosongan Database")
        dialog.setIcon(QMessageBox.Question)
        dialog.setText(
            "Sekarang telah memasuki awal bulan. Setiap tanggal 1 awal bulan, database akan dikosongkan.\n\n"
            "Apakah Anda ingin melanjutkan?"
        )
        dialog.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        response = dialog.exec()

        if response == QMessageBox.Yes:
            # Proses menyimpan data
            save_dialog = QMessageBox()
            save_dialog.setWindowTitle("Simpan Data Terakhir")
            save_dialog.setText("Simpan data terakhir sebelum pengosongan database.")
            save_dialog.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            save_response = save_dialog.exec()

            if save_response == QMessageBox.Yes:
                # Simpan data terakhir ke Excel
                self.download_table_to_excel()

                # Dialog kedua: Konfirmasi penghapusan database
                confirm_dialog = QMessageBox()
                confirm_dialog.setWindowTitle("Konfirmasi Penghapusan")
                confirm_dialog.setText("Yakin ingin menghapus seluruh data dari database?")
                confirm_dialog.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                confirm_response = confirm_dialog.exec()

                if confirm_response == QMessageBox.Yes:
                    # Proses penghapusan data
                    self.delete_all_collections()
                    QMessageBox.information(None, "Berhasil", "Database berhasil dikosongkan.")
                    self.load_table_data()
                else:
                    QMessageBox.information(None, "Batal", "Proses penghapusan database dibatalkan.")
            else:
                QMessageBox.information(None, "Batal", "Penyimpanan data dibatalkan. Proses dihentikan.")
        else:
            QMessageBox.information(None, "Batal", "Proses pengosongan database dibatalkan.")

    def check_date_and_clear_database_once(self, debug_date=None):
        """Memeriksa apakah dialog pengosongan database perlu ditampilkan.
        Args:
            debug_date (QDate, optional): Tanggal khusus untuk debugging.
        """
        # Tentukan lokasi file untuk menyimpan bulan terakhir
        last_cleared_file = f"databases/last_cleared_month.txt"
        
        self.resource_path(last_cleared_file)

        # Ambil tanggal saat ini, atau gunakan debug_date jika diberikan
        current_date = debug_date if debug_date else QDate.currentDate()
        if not isinstance(current_date, QDate):
            raise ValueError("current_date harus berupa QDate")

        # Format tahun dan bulan untuk identifikasi unik
        current_year_month = f"{current_date.year()}-{current_date.month()}"

        # Periksa apakah file bulan terakhir ada
        if os.path.exists(last_cleared_file):
            with open(last_cleared_file, "r") as file:
                last_cleared_month = file.read().strip()
        else:
            last_cleared_month = None

        # Jika bulan saat ini berbeda dari bulan terakhir, tampilkan dialog
        if last_cleared_month != current_year_month:
            self.show_clear_database_dialog()

            # Simpan bulan saat ini sebagai bulan terakhir
            with open(last_cleared_file, "w") as file:
                file.write(current_year_month)
            
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    sys.exit(app.exec())